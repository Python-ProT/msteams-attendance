import numpy as np
import pandas as pd
import pymongo
import openpyxl
from openpyxl.styles import PatternFill

from flask import Flask, flash, request, redirect, url_for, render_template,send_file
# import urllib.request
import os
from os.path import join, dirname, realpath
from werkzeug.utils import secure_filename
from dotenv import load_dotenv

load_dotenv()
# API_URL=os.getenv("API_URL")
API_URL="mongodb+srv://project-prot:9557930603@cluster1.nan0h.mongodb.net/test"
app = Flask(__name__)
 
# UPLOADS_PATH = 'static/uploads/'
UPLOADS_PATH = join(dirname(realpath(__file__)), 'static/uploads')
 
app.secret_key = "secret key"
app.config['UPLOADS_PATH'] = UPLOADS_PATH
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
 
# ALLOWED_EXTENSIONS = set(['png', 'xlsx'])

# ALLOWED_EXTENSIONS = set(['txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif', 'csv','xlsx'])
ALLOWED_EXTENSIONS = set(['csv','xlsx'])

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
     
 
@app.route('/')
def home():
    return render_template('index.html')


@app.route('/submit',methods=['POST','GET'])
def submit():

    # os.remove(os.path.join(app.config['UPLOADS_PATH'], filename2))  
    dir = 'static/uploads'
    for f in os.listdir(dir):
        os.remove(os.path.join(dir, f))
    if 'files[]' not in request.files:
        flash('No file Selected') 
        return redirect(request.url)
    # file = request.files['file']
    files = request.files.getlist('files[]')
    file2 = request.files['file2']
    print(len(files))
    print(files)
    # if file2.filename == '':
    #     flash('No File selected for uploading')
    #     return redirect(request.url)
    if file2 and allowed_file(file2.filename):
        # filename = secure_filename(file.filename)
        global filename2
        filename2 = secure_filename(file2.filename)
        
        # file.save(os.path.join(app.config['UPLOADS_PATH'], filename))
        file2.save(os.path.join(app.config['UPLOADS_PATH'], filename2))
        for file in files:
            print(file)
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                file.save(os.path.join(app.config['UPLOADS_PATH'], filename))
                print(os.path.join(app.config['UPLOADS_PATH'], filename))
            compute(filename,filename2)
            os.remove(os.path.join(app.config['UPLOADS_PATH'], filename))
        # download(filename2)
        
        return render_template('index.html')
    else:
        return redirect(request.url) 

@app.route('/download')
def download():
    global filename2

    print("Welcome to pymongo")
    client = pymongo.MongoClient(API_URL)
    print(client)

    # df1.to_csv("try.csv", index=None, header=True)
    # df = pd.read_csv(df1)
    db =client["Attendance"]
    print(db)
    collection=db[filename2]

    os.remove(os.path.join(app.config['UPLOADS_PATH'], filename2))
    all_docs = collection.find({},{'_id':0})
    list_cursor=list(all_docs)
    df3=pd.DataFrame(list_cursor)
    df3=df3.set_index("Scholar No")
    # print(df3)
    df3.to_excel(os.path.join(app.config['UPLOADS_PATH'], filename2))
    wb = openpyxl.load_workbook(os.path.join(app.config['UPLOADS_PATH'], filename2))
    ws = wb['Sheet1']
    fill_pattern = PatternFill(patternType='solid', fgColor='C64747')

    for j in range(0, len(df3["Full Name"])):
        # print(type(df3["Percentage"][j]))
        s=float(df3["Percentage"][j])
        # print(type(s))

        if(s < 75.0):
            my_list = list(df3)
            index = my_list.index("Percentage")
            col = chr(index+65+1)
            ws[col+str(j+2)].fill = fill_pattern
            wb.save(os.path.join(app.config['UPLOADS_PATH'], filename2))
    uploads=os.path.join(app.config['UPLOADS_PATH'], filename2)
    # print(uploads)

    return send_file(uploads,as_attachment=True) 



# 09:15:00 AM
def difftime(a, b):
    a = a.split(":")
    b = b.split(":")
    if b[-1][-2].upper() == "P" and (int)(b[0]) != 12:
        b[0] = str(int(b[0]) + 12)
    if a[-1][-2].upper() == "P" and (int)(a[0]) != 12:
        a[0] = str(int(a[0]) + 12)
    return (int(b[0]) - int(a[0]))*3600 + (int(b[1]) - int(a[1]))*60 + int(b[2][:2]) - int(a[2][:2])


# Minimum time calculation for setting start and end time for individual students
def minitime(a, b):
    a = a.split(":")
    b = b.split(":")
    if b[-1][-2].upper() == "P" and a[-1][-2].upper() == "A":
        return True
    elif (b[-1][-2].upper() == "P" and a[-1][-2].upper() == "P") or (b[-1][-2].upper() == "A" and a[-1][-2].upper() == "A"):
        if int(a[0]) < int(b[0]):
            return True
        elif int(a[0]) == int(b[0]):
            if int(a[1]) < int(b[1]):
                return True
            elif int(a[1]) == int(a[1]):
                if int(a[2][:2]) < int(b[2][:2]) or int(a[2][:2]) == int(b[2][:2]):
                    return True
                else:
                    return False
            else:
                return False
        else:
            return False
    else:
        return False
def compute(filename,filename2):
    totol = 0
    file2 = os.path.join(app.config['UPLOADS_PATH'], filename)
    file1 = os.path.join(app.config['UPLOADS_PATH'], filename2)
    df1 = pd.read_excel(file1)
    # print(df1)
    print("Welcome to pymongo")
    client = pymongo.MongoClient(API_URL)
    print(client)

    # df1.to_csv("try.csv", index=None, header=True)
    # df = pd.read_csv(df1)
    db =client["Attendance"]
    # print(db)
    collection=db[filename2]
    data = df1.to_dict(orient="records")
    # print(data)
    if collection.count()==0:
        collection.insert_many(data)
        
    # path = os.path.join(app.config['UPLOADS_PATH'], filename2)

    df2 = pd.read_csv(file2, encoding="utf-16", sep='\t')
    # print(df2)
    newdate = df2["Timestamp"][0].split(',')[0]
    date = newdate.split('/')[1] + "/" + \
        newdate.split('/')[0] + "/" + newdate.split('/')[2]
    # print(date)
    # print(df2["Timestamp"][0].split(',')[0])
    start=request.form['start']
    end=request.form['end']
    start = start.split(":")      
    end = end.split(":")
    
    t ="AM"
    if start[0]>"12":
        start[0]=int(start[0])-12
        t="PM"
    

    h =str(start[0])
    
    m =start[1]
    s ="00"
    # t =request.form['clicked4']
    if len(h) == 1:
        h = "0" + str(h)
    if len(m) == 1:
        m = "0" + str(m)
    if len(s) == 1:
        s = "0" + str(s)
    start_time = h + ":" + m + ":" + s + " " + t
    print(start_time)

    if end[0]>"12":
        end[0]=int(end[0])-12
        t="PM" 
    h =str(end[0])
    m = end[1]
    s = "00"
    # t = request.form['clicked8']
    if len(h) == 1:
        h = "0" + str(h)
    if len(m) == 1:
        m = "0" + str(m)
    if len(s) == 1:
        s = "0" + str(s)
    end_time = h + ":" + m + ":" + s + " " + t
    # print(end_time)

    thers = "25"
    # print(thers)
    # print(date)
    # print(type(df2["Full Name"]))
    stud_attend = {}.fromkeys(df2["Full Name"])
    # print(stud_attend)
    for i in stud_attend:
        i.strip(" ").upper()
        # print(i.strip(" ").upper())
    # Timestamp extraction for individual students
    l = []
    # ye ek ka time bta rha h ek jagah pe
    for i in stud_attend:
        for j in range(0, len(df2["Full Name"])):
            if i == df2["Full Name"][j]:
                l.append(df2["Timestamp"][j].split(",")[1].lstrip(" "))
                # print(l)
        stud_attend[i] = l
        l = []
        # print(l)
    # Calculate difference between every joining and left time and calculate whether the person is present for given threshold
    # True if he is present and False, if he is absent
    tot_time = difftime(start_time, end_time)
    print(tot_time)
    # If the person not attended meeting itself means set False for that person
    for i in stud_attend:
        sumtime = 0
        x = stud_attend[i]
        if minitime(x[0], start_time):
            x[0] = start_time
        if (len(x) % 2) == 0 and minitime(end_time, x[-1]):
            x[-1] = end_time
        if (len(x) % 2) != 0:
            x.append(end_time)
            # print(x.append(end_time))
        for j in range(0, len(x), 2):
            sum_time = difftime(x[j], x[j+1])
            if sum_time >= eval(thers)*tot_time / 100:
                stud_attend[i] = "P"
            else:
                stud_attend[i] = "A"




    # Empty column creation for the person to append the attendance list created
    for i in df1["Full Name"]:
        # print(i.strip(" ").upper())
        if i.strip(" ").upper() not in stud_attend:
            stud_attend[i.strip(" ").upper()] = "A"
    # Empty column creation for the person to append the attendance list created


    # For each person put the value True or False according to rules defined above
    d = list(df1["Full Name"])

    

    all_docs = collection.find({},{'_id':0})
    list_cursor=list(all_docs)
    df5=pd.DataFrame(list_cursor)
    # print(df5)
    datenorep=collection.find({date:{"$exists":True}})
    if datenorep=="":
        df5[date] = list(range(0, len(df5["Full Name"])))
        for i in d:
            df5[date][d.index(i)] = stud_attend[i.strip(" ").upper()]
        # for j in range(0, len(df1["Full Name"])):
        #     prev = {'Full Name':df1['Full Name'][j]}
        #     nextt3={'$set':{date:str(df1[date][j])}}
        #     collection.update_one(prev,nextt3)
        #     print(stud_attend[i.strip(" ").upper()])

        
        df1 = df1.set_index("Scholar No")
        df1 = df1.sort_values("Scholar No")


        totol=0
        df5 = df5.set_index("Scholar No")
        if "Total" not in df5:
            df5["Total"] = 0

        if "Percentage" not in df5:
            df5["Percentage"] = 0
        for column in df5.columns:
            if ((column != "Full Name") and (column != "Total") and (column != "Percentage")):
                totol += 1

        for j in range(0, len(df5["Full Name"])):
            # print(df3[date][j]," ",df3["Tota"][j])
            py_int = np.int64(df5["Total"][j]).item()
            if df5[date][j] == "P":
                py_int += 1
                df5["Total"][j] = py_int
            df5["Percentage"][j] = py_int*100/totol

            df3 = df5[["Full Name"]].copy()
        for column in df5.columns:
            if ((column != "Scholar No") or (column != "Full Name") or (column != "Total") or (column != "Percentage")):
                df3[column] = df5[column]
        

        for j in range(0, len(df5["Full Name"])):
            prev = {'Full Name':df5['Full Name'][j]}
            print(df5['Total'][j])
            nextt = {'$set':{'Total':str(df5['Total'][j])}}
            nextt2={'$set':{'Percentage':str(df5['Percentage'][j])}}
            nextt3={'$set':{date:str(df3[date][j])}}
            collection.update_one(prev,nextt)
            collection.update_one(prev,nextt2)
            collection.update_one(prev,nextt3)


    else:
        print("Already exist")


    print('\nDone!')

    




if __name__=="__main__":
    app.run(debug=True)
